from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from app.database.connection import (
    SessionLocal
)

from app.database.models import (
    Invitation
)


def confirm_invitation(
    name: str
):

    db = SessionLocal()

    guest = name.strip().title()

    try:

        invitation = (
            db.query(Invitation)
            .filter(
                Invitation.name == guest
            )
            .first()
        )

        if not invitation:

            return {
                "success": False,
                "error": "Convite não encontrado"
            }

        if invitation.confirmed:

            return {
                "success": False,
                "error": "Presença já confirmada"
            }

        invitation.confirmed = True

        db.commit()

        return {
            "success": True,
            "message": "Presença confirmada"
        }

    finally:

        db.close()


async def import_invitations_from_excel(
    file
):

    db = SessionLocal()

    if not file.filename.lower().endswith(
        ".xlsx"
    ):
        return {
            "success": False,
            "error": "Arquivo deve ser .xlsx"
        }

    try:

        with NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as temp_file:

            content = await file.read()

            temp_file.write(content)

            temp_file.flush()

            temp_path = temp_file.name

        workbook = load_workbook(
            temp_path
        )

        sheet = workbook.active

        imported = 0
        skipped = 0

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if not row:
                continue

            name = row[0]

            if not name:
                continue

            guest = str(
                name
            ).strip().title()

            existing_invitation = (
                db.query(Invitation)
                .filter(
                    Invitation.name == guest
                )
                .first()
            )

            if existing_invitation:

                skipped += 1

                continue

            invitation = Invitation(
                name=guest,
                confirmed=False
            )

            db.add(invitation)

            imported += 1

        db.commit()

        return {
            "success": True,
            "imported": imported,
            "skipped": skipped
        }

    finally:

        db.close()
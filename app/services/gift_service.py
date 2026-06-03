from app.database.connection import SessionLocal
from app.database.models import Gift

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile


def get_gifts():

    db = SessionLocal()

    try:

        return db.query(Gift).all()

    finally:

        db.close()


def create_gift(
    name: str,
    value: float,
    image: str | None = None
):

    db = SessionLocal()

    try:

        gift = Gift(
            name=name,
            value=value,
            image=(
                f'https://ipmpmkevpsabdhbeotxc.supabase.co/storage/v1/object/public/gift-images/{image}'
                if image
                else None
            )
        )

        db.add(gift)

        db.commit()

        db.refresh(gift)

        return gift

    finally:

        db.close()


async def import_gifts_from_excel(file):

    db = SessionLocal()

    if not file.filename.endswith(".xlsx"):
        return {
            "success": False,
            "error": "Arquivo deve ser .xlsx"
    }

    try:

        with NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as temp_file:

            content = await file.read()

            temp_file.write(content)

            workbook = load_workbook(
                temp_file.name
            )

        sheet = workbook.active

        imported = 0

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            name, value, image = row

            if not name:
                continue

            existing_gift = (
                db.query(Gift)
                .filter(Gift.name == name)
                .first()
            )

            if existing_gift:
                continue

            gift = Gift(
                name=name,
                value=float(value),
                image=(
                    f"https://ipmpmkevpsabdhbeotxc.supabase.co/storage/v1/object/public/gift-images/{image}"
                    if image
                    else None
                )
            )

            db.add(gift)

            imported += 1

        db.commit()

        return {
            "success": True,
            "imported": f'{imported} produtos foram importados'
        }

    finally:

        db.close()